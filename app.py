# ============================================================
#  B2B Revenue Reverse Funnel Planner
#  Methodology: PERT Estimation + Monte Carlo Simulation
#  © Marko Gross
#
#  Run:  streamlit run app.py
# ============================================================

import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import io
from datetime import datetime

# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="B2B Reverse Funnel Planner · Marko Gross",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    /* ── Metrics ───────────────────────────────────────────── */
    div[data-testid="metric-container"] {
        background-color: var(--secondary-background-color);
        border: 1px solid rgba(128,128,128,0.2);
        border-radius: 8px;
        padding: 12px 16px;
    }

    /* ── Sidebar ───────────────────────────────────────────── */
    div[data-testid="stSidebar"] { background-color: #0d1b2a; }
    div[data-testid="stSidebar"] * { color: #e0e0e0 !important; }

    /* ── Headings ──────────────────────────────────────────── */
    .section-header {
        font-size: 1rem; font-weight: 700;
        color: #0066cc; margin-bottom: 4px;
        border-bottom: 2px solid rgba(128,128,128,0.2);
        padding-bottom: 4px;
    }

    /* ── Tab bar container ─────────────────────────────────── */
    div[data-testid="stTabs"] > div:first-child {
        background-color: var(--secondary-background-color);
        border-radius: 10px 10px 0 0;
        padding: 4px 6px 0 6px;
        border-bottom: 3px solid #0066cc;
        gap: 4px;
    }

    /* ── All tab buttons ───────────────────────────────────── */
    button[data-baseweb="tab"] {
        font-size: 0.88rem !important;
        font-weight: 600 !important;
        color: var(--text-color) !important;
        background-color: var(--secondary-background-color) !important;
        border-radius: 8px 8px 0 0 !important;
        padding: 10px 18px !important;
        border: 1px solid rgba(128,128,128,0.25) !important;
        border-bottom: none !important;
        margin-right: 2px !important;
        transition: background-color 0.15s, color 0.15s !important;
    }

    /* ── Hover state ───────────────────────────────────────── */
    button[data-baseweb="tab"]:hover {
        background-color: rgba(0,102,204,0.15) !important;
        color: #0066cc !important;
    }

    /* ── Active / selected tab ─────────────────────────────── */
    button[data-baseweb="tab"][aria-selected="true"] {
        background-color: #0066cc !important;
        color: #ffffff !important;
        border-color: #0066cc !important;
        font-weight: 700 !important;
    }

    /* ── Remove the default Streamlit underline indicator ─── */
    button[data-baseweb="tab"][aria-selected="true"]::after,
    button[data-baseweb="tab"]::after { display: none !important; }

    /* ── Docs: theme-aware boxes ───────────────────────────── */
    .docs-de {
        background: rgba(0,102,204,0.08);
        border-left: 3px solid #0066cc;
        padding: 10px 14px; border-radius: 4px; margin-bottom: 6px;
        color: var(--text-color);
    }
    .docs-en {
        background: var(--secondary-background-color);
        border-left: 3px solid rgba(128,128,128,0.4);
        padding: 10px 14px; border-radius: 4px; margin-bottom: 12px;
        color: var(--text-color);
    }
    .docs-label-de {
        font-size: 0.72rem; font-weight: 700; color: #0066cc;
        text-transform: uppercase; letter-spacing: .05em;
    }
    .docs-label-en {
        font-size: 0.72rem; font-weight: 700;
        color: var(--text-color); opacity: 0.55;
        text-transform: uppercase; letter-spacing: .05em;
    }

    /* ── Settings bar & custom boxes: theme-aware ──────────── */
    .theme-box {
        background: var(--secondary-background-color);
        border-left: 4px solid #0066cc;
        color: var(--text-color);
        padding: 7px 16px; border-radius: 4px; margin-bottom: 14px;
        font-size: 0.82rem; display: flex; gap: 20px;
        flex-wrap: wrap; align-items: center;
    }
    .theme-box span { color: var(--text-color) !important; }
    .theme-box .accent { color: #0066cc !important; }
    .theme-box .muted { opacity: 0.55; font-size: 0.74rem; margin-left: auto; }
    .theme-card {
        background: var(--secondary-background-color);
        border-radius: 6px; border: 1px solid rgba(128,128,128,0.2);
        color: var(--text-color); padding: 10px 14px; margin-top: 16px;
    }
    .theme-warn {
        background: rgba(255,193,7,0.12);
        border-left: 3px solid #ffc107;
        padding: 12px 16px; border-radius: 4px; margin-bottom: 12px;
        color: var(--text-color);
    }
</style>
""", unsafe_allow_html=True)

# ============================================================
# CONSTANTS
# ============================================================

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

FUNNEL_ARCHETYPES = {
    "Classic B2B": {
        "labels":      ["Win Rate (Opp→Deal)", "SQL→Opp", "MQL→SQL", "Lead→MQL", "Touchpoint→Lead"],
        "stage_names": ["Revenue", "Deals", "Opportunities", "SQLs", "MQLs", "Leads", "Touchpoints"],
        "base":  [25, 40, 30, 15, 2],
        "worst": [15, 25, 20, 10, 1],
        "best":  [35, 55, 40, 20, 3],
    },
    "Enterprise": {
        "labels":      ["Win Rate (Opp→Deal)", "SQL→Opp", "MQL→SQL", "Lead→MQL", "Touchpoint→Lead"],
        "stage_names": ["Revenue", "Deals", "Opportunities", "SQLs", "MQLs", "Leads", "Touchpoints"],
        "base":  [18, 35, 60, 25, 1],
        "worst": [12, 20, 45, 15, 1],
        "best":  [25, 50, 75, 35, 2],
    },
    "SaaS / PLG": {
        "labels":      ["Win Rate (Opp→Deal)", "SQL→Opp", "PQL→SQL", "Signup→PQL", "Visitor→Signup"],
        "stage_names": ["Revenue", "Deals", "Opportunities", "PQLs", "Signups", "Visitors", "Touchpoints"],
        "base":  [20, 45, 40, 15, 5],
        "worst": [12, 30, 25,  8, 2],
        "best":  [30, 60, 55, 25, 10],
    },
    "Channel / Partner": {
        "labels":      ["Win Rate (Opp→Deal)", "Partner SQL→Opp", "Partner Lead→SQL", "Contact→Lead", "Touchpoint→Contact"],
        "stage_names": ["Revenue", "Deals", "Opportunities", "SQLs", "MQLs", "Contacts", "Touchpoints"],
        "base":  [30, 50, 35, 20, 4],
        "worst": [20, 35, 20, 10, 2],
        "best":  [40, 70, 50, 30, 7],
    },
}

SEASONALITY_PROFILES = {
    "Flat":            [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
    "Q4 heavy":        [0.7, 0.8, 0.9, 1.0, 1.0, 1.0, 0.8, 0.8, 1.1, 1.3, 1.4, 1.2],
    "Events & Autumn": [1.1, 1.1, 1.2, 1.0, 0.9, 0.8, 0.7, 0.8, 1.2, 1.3, 1.2, 0.7],
    "Q1 Budget Push":  [1.3, 1.2, 1.1, 0.9, 0.8, 0.8, 0.7, 0.7, 0.9, 1.0, 1.0, 1.0],
    "Summer Slowdown": [1.0, 1.0, 1.0, 1.1, 1.1, 0.9, 0.6, 0.6, 1.0, 1.2, 1.3, 1.2],
    "Mid-year Ramp":   [0.8, 0.9, 1.0, 1.1, 1.2, 1.3, 1.2, 1.1, 1.0, 0.9, 0.8, 0.8],
    "Spring Launch":   [1.1, 1.2, 1.3, 1.2, 1.0, 0.8, 0.7, 0.7, 0.8, 0.9, 0.9, 0.8],
    "Autumn Launch":   [0.8, 0.8, 0.9, 1.0, 1.0, 0.9, 0.9, 1.0, 1.2, 1.3, 1.3, 0.9],
    "Eigenes Profil":  [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
}

DEFAULT_CHANNELS = [
    {"group": "Paid Search",          "activity": "Brand Campaigns",         "cost_per_mql": 40,  "share": 14.0},
    {"group": "Paid Search",          "activity": "Non-Brand Campaigns",     "cost_per_mql": 70,  "share": 12.0},
    {"group": "Paid Social",          "activity": "Prospecting",             "cost_per_mql": 60,  "share": 11.0},
    {"group": "Paid Social",          "activity": "Retargeting",             "cost_per_mql": 40,  "share":  8.0},
    {"group": "Website",              "activity": "Organic Onsite",          "cost_per_mql": 15,  "share":  6.0},
    {"group": "Website",              "activity": "Direct & Referral",       "cost_per_mql": 15,  "share":  6.0},
    {"group": "Events",               "activity": "Trade Shows",             "cost_per_mql": 400, "share":  3.0},
    {"group": "Events",               "activity": "Roundtables & Sponsoring","cost_per_mql": 500, "share":  4.0},
    {"group": "Content",              "activity": "Blogs & Articles",        "cost_per_mql": 20,  "share":  5.0},
    {"group": "Content",              "activity": "Whitepapers",             "cost_per_mql": 60,  "share":  4.0},
    {"group": "Content",              "activity": "Case Studies",            "cost_per_mql": 100, "share":  3.0},
    {"group": "SEO & AI Visibility",  "activity": "Technical SEO",           "cost_per_mql": 15,  "share":  5.0},
    {"group": "SEO & AI Visibility",  "activity": "Content Cluster SEO",     "cost_per_mql": 25,  "share":  5.0},
    {"group": "ABM",                  "activity": "1:1 Enterprise ABM",      "cost_per_mql": 600, "share":  2.0},
    {"group": "ABM",                  "activity": "1:Few ABM",               "cost_per_mql": 200, "share":  3.0},
    {"group": "Marketing Automation", "activity": "Nurture Programs",        "cost_per_mql": 10,  "share":  5.0},
    {"group": "Marketing Automation", "activity": "Reactivation Flows",      "cost_per_mql": 15,  "share":  4.0},
]

# ============================================================
# HELPERS
# ============================================================

def make_cr_df(archetype: str) -> pd.DataFrame:
    a = FUNNEL_ARCHETYPES[archetype]
    return pd.DataFrame({
        "Stage":      a["labels"],
        "Worst (%)":  a["worst"],
        "Base (%)":   a["base"],
        "Best (%)":   a["best"],
    })

def run_funnel(revenue_target, deal_size, cr_pct):
    crs = [c / 100 for c in cr_pct]
    deals       = revenue_target / deal_size
    opps        = deals  / crs[0]
    stage3      = opps   / crs[1]
    stage4      = stage3 / crs[2]
    stage5      = stage4 / crs[3]
    touchpoints = stage5 / crs[4]
    return deals, opps, stage3, stage4, stage5, touchpoints

def calc_channel_budget(mqls, ch_df):
    df = ch_df.copy()
    total = df["share"].sum() or 1
    df["share_norm"]     = df["share"] / total
    df["planned_mqls"]   = df["share_norm"] * mqls
    df["required_spend"] = df["planned_mqls"] * df["cost_per_mql"]
    return df

@st.cache_data(show_spinner=False)
def run_monte_carlo(revenue_target, deal_size, cr_means_pct, worst_pct, best_pct,
                    avg_cpm, n_sims, seed=42):
    cr_means = [c / 100 for c in cr_means_pct]
    cr_stds  = [(b - w) / 600 for w, b in zip(worst_pct, best_pct)]
    rng = np.random.default_rng(seed)
    results = []
    for _ in range(n_sims):
        crs = [float(np.clip(rng.normal(m, s), 0.005, 0.99))
               for m, s in zip(cr_means, cr_stds)]
        cpm = max(1.0, float(rng.normal(avg_cpm, avg_cpm * 0.25)))
        deals  = revenue_target / deal_size
        opps   = deals  / crs[0]
        s3     = opps   / crs[1]
        s4     = s3     / crs[2]
        s5     = s4     / crs[3]
        tp     = s5     / crs[4]
        results.append({"deals": deals, "opps": opps,
                        "stage3": s3, "stage4": s4, "stage5": s5,
                        "touchpoints": tp, "budget_req": s4 * cpm})
    return pd.DataFrame(results)

def get_season_weights(profile):
    idx   = SEASONALITY_PROFILES[profile]
    total = sum(idx)
    return [v / total for v in idx]

def _settings_bar(archetype, scenario, season_profile, coverage):
    """#15 — Compact active-settings info row shown at the top of every results tab."""
    cov_icon = "🟢" if coverage >= 0.9 else ("🟡" if coverage >= 0.6 else "🔴")
    scen_badge = {"Worst": "#dc3545", "Base": "#0066cc", "Best": "#28a745"}.get(scenario, "#0066cc")
    st.markdown(
        f"""<div class="theme-box">
        <span style="font-weight:600;">🏗️ Archetype: <span class="accent">{archetype}</span></span>
        <span style="font-weight:600;">📊 Szenario:
          <span style="background:{scen_badge};color:#fff;padding:1px 7px;border-radius:10px;font-size:0.78rem;">{scenario}</span>
        </span>
        <span style="font-weight:600;">📅 Saisonalität: {season_profile}</span>
        <span style="font-weight:600;">{cov_icon} Coverage: {coverage:.0%}</span>
        <span class="muted">← ⚙️ Inputs-Tab zum Ändern</span>
        </div>""",
        unsafe_allow_html=True,
    )

# ============================================================
# EXPORT HELPERS
# ============================================================

def generate_excel(revenue_target, deal_size, available_budget, archetype, scenario,
                   stage_names, funnel_vals, stage4, deals, total_required, coverage,
                   plan_mqls, plan_deals, plan_revenue, plan_budget,
                   actual_df, ch_calc):
    """Build an in-memory .xlsx and return bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="0066CC")
    ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(size=10)
    TITLE_FONT= Font(bold=True, size=12, color="1A1A2E")
    THIN      = Side(style="thin", color="D0D0D0")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")

    def hdr_row(ws, row, values, widths=None):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        if widths:
            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

    def data_row(ws, row, values, formats=None, alt=False):
        fill = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if formats and c <= len(formats) and formats[c-1]:
                cell.number_format = formats[c-1]

    # ── Sheet 1: Funnel Summary ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Funnel Summary"
    ws1.row_dimensions[1].height = 24

    ws1.cell(1, 1, "B2B Revenue Reverse Funnel — Summary").font = TITLE_FONT
    ws1.cell(2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} · © Marko Gross")
    ws1.cell(2, 1).font = Font(italic=True, size=9, color="888888")
    ws1.merge_cells("A1:D1")
    ws1.merge_cells("A2:D2")

    ws1.cell(4, 1, "Parameter").font = Font(bold=True)
    ws1.cell(4, 2, "Value").font = Font(bold=True)
    params = [
        ("Revenue Target",        revenue_target,    '€#,##0'),
        ("Avg. Deal Size",        deal_size,         '€#,##0'),
        ("Available Budget",      available_budget,  '€#,##0'),
        ("Required Budget",       total_required,    '€#,##0'),
        ("Budget Coverage",       coverage,          '0%'),
        ("Deals needed",          deals,             '#,##0.0'),
        ("MQLs needed",           stage4,            '#,##0.0'),
        ("Funnel Archetype",      archetype,         None),
        ("Scenario",              scenario,          None),
    ]
    for r, (label, val, fmt) in enumerate(params, 5):
        c1 = ws1.cell(r, 1, label);  c1.font = BODY_FONT
        c2 = ws1.cell(r, 2, val);    c2.font = BODY_FONT
        if fmt: c2.number_format = fmt
        f = ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        c1.fill = c2.fill = f

    ws1.cell(15, 1, "Funnel Stages").font = Font(bold=True)
    hdr_row(ws1, 16, ["Stage", "Value"], [28, 20])
    for r, (name, val) in enumerate(zip(stage_names, funnel_vals), 17):
        fmt = '€#,##0' if r == 17 else '#,##0.0'
        data_row(ws1, r, [name, val], [None, fmt], alt=(r % 2 == 0))

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Monthly Plan ────────────────────────────────
    ws2 = wb.create_sheet("Monthly Plan")
    hdr_row(ws2, 1,
            ["Month", "Plan MQLs", "Plan Deals", "Plan Revenue (€)", "Plan Budget (€)"],
            [10, 14, 14, 20, 18])
    for r, m in enumerate(MONTHS, 2):
        data_row(ws2, r,
                 [m, plan_mqls[r-2], plan_deals[r-2], plan_revenue[r-2], plan_budget[r-2]],
                 [None, '#,##0.0', '#,##0.0', '€#,##0', '€#,##0'],
                 alt=(r % 2 == 0))
    totals_r = 14
    data_row(ws2, totals_r,
             ["TOTAL",
              sum(plan_mqls), sum(plan_deals),
              sum(plan_revenue), sum(plan_budget)],
             [None, '#,##0.0', '#,##0.0', '€#,##0', '€#,##0'])
    ws2.cell(totals_r, 1).font = Font(bold=True)

    # ── Sheet 3: Plan vs. Actual ────────────────────────────
    ws3 = wb.create_sheet("Plan vs. Actual")
    hdr_row(ws3, 1,
            ["Month",
             "Plan MQLs", "Actual MQLs", "Δ MQLs",
             "Plan Revenue (€)", "Actual Revenue (€)", "Δ Revenue (€)"],
            [10, 14, 14, 12, 18, 20, 16])
    adf = actual_df.reset_index(drop=True)
    for r, m in enumerate(MONTHS, 2):
        pm = plan_mqls[r-2]; am = float(adf.loc[r-2, "Actual MQLs"])
        pr = plan_revenue[r-2]; ar = float(adf.loc[r-2, "Actual Revenue (€)"])
        data_row(ws3, r,
                 [m, pm, am, am - pm, pr, ar, ar - pr],
                 [None, '#,##0.0', '#,##0.0', '+#,##0.0;-#,##0.0',
                  '€#,##0', '€#,##0', '+€#,##0;-€#,##0'],
                 alt=(r % 2 == 0))

    # ── Sheet 4: Channels ───────────────────────────────────
    ws4 = wb.create_sheet("Channels")
    hdr_row(ws4, 1,
            ["Group", "Activity", "Cost/MQL (€)", "Share (%)", "Planned MQLs", "Required Spend (€)"],
            [22, 28, 14, 12, 16, 20])
    for r, (_, row) in enumerate(ch_calc.iterrows(), 2):
        data_row(ws4, r,
                 [row["group"], row["activity"],
                  row["cost_per_mql"], row["share_norm"] * 100,
                  row["planned_mqls"], row["required_spend"]],
                 [None, None, '€#,##0', '#,##0.0"%"', '#,##0.0', '€#,##0'],
                 alt=(r % 2 == 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_pdf(revenue_target, deal_size, available_budget, archetype, scenario,
                 stage_names, funnel_vals, stage4, deals, total_required, coverage,
                 plan_mqls, plan_revenue, ch_calc):
    """Build an in-memory PDF and return bytes."""
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    BLUE  = (0,   102, 204)
    DARK  = (26,  26,  46)
    LGRAY = (235, 243, 251)
    WHITE = (255, 255, 255)
    DGRAY = (100, 100, 100)

    class PDF(FPDF):
        def header(self):
            self.set_fill_color(*BLUE)
            self.rect(0, 0, 210, 18, "F")
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(*WHITE)
            self.set_xy(8, 3)
            self.cell(0, 12, "B2B Revenue Reverse Funnel Planner", ln=False)
            self.set_font("Helvetica", "", 8)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            self.set_xy(-60, 5)
            self.cell(52, 8, ts, align="R")
            self.ln(20)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(*DGRAY)
            self.cell(0, 6, f"© Marko Gross  ·  B2B Revenue Reverse Funnel Planner  ·  Page {self.page_no()}", align="C")

    pdf = PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    def section_title(title):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*BLUE)
        pdf.set_fill_color(*LGRAY)
        pdf.cell(0, 7, f"  {title}", ln=True, fill=True)
        pdf.set_text_color(*DARK)
        pdf.ln(1)

    def kv_row(label, value, alt=False):
        pdf.set_font("Helvetica", "", 9)
        if alt:
            pdf.set_fill_color(*LGRAY)
        else:
            pdf.set_fill_color(*WHITE)
        pdf.set_text_color(*DARK)
        pdf.cell(70, 6, label, fill=True, border=0)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 6, str(value), ln=True, fill=True, border=0)

    # ── Key Metrics ──────────────────────────────────────────
    section_title("Key Metrics")
    kv_row("Revenue Target",   f"EUR {revenue_target:,.0f}",  alt=False)
    kv_row("Avg. Deal Size",   f"EUR {deal_size:,.0f}",       alt=True)
    kv_row("Available Budget", f"EUR {available_budget:,.0f}",alt=False)
    kv_row("Required Budget",  f"EUR {total_required:,.0f}",  alt=True)
    kv_row("Budget Coverage",  f"{coverage:.0%}",          alt=False)
    kv_row("Deals needed",     f"{deals:,.0f}",            alt=True)
    kv_row("MQLs needed",      f"{stage4:,.0f}",           alt=False)
    kv_row("Funnel Archetype", archetype,                  alt=True)
    kv_row("Scenario",         scenario,                   alt=False)
    pdf.ln(4)

    # ── Funnel Stages ────────────────────────────────────────
    section_title("Funnel Stages")
    col_w = [80, 40]
    pdf.set_fill_color(*BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w[0], 6, "Stage",  fill=True, border=1)
    pdf.cell(col_w[1], 6, "Value",  fill=True, border=1, ln=True)
    for i, (name, val) in enumerate(zip(stage_names, funnel_vals)):
        pdf.set_fill_color(*(LGRAY if i % 2 == 0 else WHITE))
        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(col_w[0], 6, name, fill=True, border=1)
        txt = f"EUR {val:,.0f}" if i == 0 else f"{val:,.0f}"
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_w[1], 6, txt, fill=True, border=1, align="R", ln=True)
    pdf.ln(4)

    # ── Monthly Revenue Plan (compact) ───────────────────────
    section_title("Monthly Revenue Plan")
    col_w2 = [18, 33, 33]
    pdf.set_fill_color(*BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(["Month", "Plan MQLs", "Plan Revenue (EUR)"], col_w2):
        pdf.cell(w, 6, h, fill=True, border=1, align="C")
    pdf.ln()
    for i, m in enumerate(MONTHS):
        pdf.set_fill_color(*(LGRAY if i % 2 == 0 else WHITE))
        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(col_w2[0], 5, m,                         fill=True, border=1, align="C")
        pdf.cell(col_w2[1], 5, f"{plan_mqls[i]:,.0f}",    fill=True, border=1, align="R")
        pdf.cell(col_w2[2], 5, f"EUR {plan_revenue[i]:,.0f}",fill=True, border=1, align="R")
        pdf.ln()
    pdf.ln(4)

    # ── Top Channels (by spend) ─────────────────────────────
    section_title("Top Channels by Required Spend")
    grp = (ch_calc.groupby("group", as_index=False)
           .agg(mqls=("planned_mqls","sum"), spend=("required_spend","sum"))
           .sort_values("spend", ascending=False).head(8))
    col_w3 = [60, 30, 40]
    pdf.set_fill_color(*BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(["Channel Group", "MQLs", "Required Spend (EUR)"], col_w3):
        pdf.cell(w, 6, h, fill=True, border=1, align="C")
    pdf.ln()
    for i, (_, row) in enumerate(grp.iterrows()):
        pdf.set_fill_color(*(LGRAY if i % 2 == 0 else WHITE))
        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(col_w3[0], 5, row["group"],            fill=True, border=1)
        pdf.cell(col_w3[1], 5, f"{row['mqls']:,.0f}",   fill=True, border=1, align="R")
        pdf.cell(col_w3[2], 5, f"EUR {row['spend']:,.0f}", fill=True, border=1, align="R")
        pdf.ln()

    return bytes(pdf.output())


# ============================================================
# SESSION STATE — DEFAULTS
# ============================================================

_DEFAULTS = {
    "inp_revenue":       1_000_000,
    "onboarding_done":   False,
    "inp_deal_size":     5_000,
    "inp_budget":        300_000,
    "inp_archetype":     "Classic B2B",
    "inp_scenario":      "Base",
    "inp_season":        "Flat",
    "inp_custom_season": [1.0] * 12,
    "inp_n_sims":        500,
    "inp_report_month":  "Apr",
}
for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

if "inp_cr_df" not in st.session_state:
    st.session_state.inp_cr_df = make_cr_df("Classic B2B")

if "channels_df" not in st.session_state:
    st.session_state.channels_df = pd.DataFrame(DEFAULT_CHANNELS)

if "actual_df" not in st.session_state:
    st.session_state.actual_df = pd.DataFrame({
        "Month":              MONTHS,
        "Actual MQLs":        [0.0] * 12,
        "Actual Deals":       [0.0] * 12,
        "Actual Revenue (€)": [0.0] * 12,
        "Actual Budget (€)":  [0.0] * 12,
    })

# ============================================================
# READ ACTIVE INPUTS FROM SESSION STATE
# ============================================================

revenue_target  = st.session_state.inp_revenue
deal_size       = st.session_state.inp_deal_size
available_budget= st.session_state.inp_budget
archetype       = st.session_state.inp_archetype
scenario        = st.session_state.inp_scenario
season_profile  = st.session_state.inp_season
n_sims          = st.session_state.inp_n_sims
arch            = FUNNEL_ARCHETYPES[archetype]

# Active CR column
scen_col = {"Worst": "Worst (%)", "Base": "Base (%)", "Best": "Best (%)"}[scenario]
cr_pct   = st.session_state.inp_cr_df[scen_col].tolist()

# Seasonality weights
if season_profile == "Eigenes Profil":
    raw_weights = st.session_state.inp_custom_season
else:
    raw_weights = SEASONALITY_PROFILES[season_profile]
total_w        = sum(raw_weights) or 1
season_weights = [v / total_w for v in raw_weights]

# ============================================================
# MAIN CALCULATIONS
# ============================================================

try:
    deals, opps, stage3, stage4, stage5, touchpoints = run_funnel(
        revenue_target, deal_size, cr_pct
    )
except (ZeroDivisionError, ValueError):
    deals = opps = stage3 = stage4 = stage5 = touchpoints = 0.0

stage_names = arch["stage_names"]
funnel_vals = [revenue_target, deals, opps, stage3, stage4, stage5, touchpoints]

ch_calc        = calc_channel_budget(stage4, st.session_state.channels_df)
total_required = ch_calc["required_spend"].sum()
avg_cpm        = (total_required / stage4) if stage4 > 0 else 50.0
coverage       = (available_budget / total_required) if total_required > 0 else 0.0
budget_gap     = available_budget - total_required

mc_df = run_monte_carlo(
    float(revenue_target), float(deal_size),
    tuple(cr_pct), tuple(arch["worst"]), tuple(arch["best"]),
    float(avg_cpm), int(n_sims),
)

plan_mqls    = [stage4         * w for w in season_weights]
plan_deals   = [deals          * w for w in season_weights]
plan_revenue = [revenue_target * w for w in season_weights]
plan_budget  = [total_required * w for w in season_weights]

# ============================================================
# SIDEBAR — compact KPI panel
# ============================================================

with st.sidebar:
    st.markdown("### 📊 B2B Funnel Planner")
    st.caption("© Marko Gross")
    st.divider()
    cov_icon = "🟢" if coverage >= 0.9 else ("🟡" if coverage >= 0.6 else "🔴")
    st.metric("Revenue Target",   f"€{revenue_target:,.0f}")
    st.metric("Required Budget",  f"€{total_required:,.0f}")
    st.metric(f"{cov_icon} Coverage", f"{coverage:.0%}")
    st.metric("MQLs needed",      f"{stage4:,.0f}")
    st.metric("Deals needed",     f"{deals:,.0f}")
    st.divider()
    st.caption(f"Archetype: **{archetype}**")
    st.caption(f"Scenario: **{scenario}**")
    st.caption(f"Seasonality: **{season_profile}**")
    st.divider()
    st.markdown("**Export**")
    _fname_date = datetime.now().strftime("%Y%m%d")

    _excel_bytes = generate_excel(
        revenue_target, deal_size, available_budget, archetype, scenario,
        stage_names, funnel_vals, stage4, deals, total_required, coverage,
        plan_mqls, plan_deals, plan_revenue, plan_budget,
        st.session_state.actual_df, ch_calc
    )
    if _excel_bytes:
        st.download_button(
            "📥 Excel exportieren",
            data=_excel_bytes,
            file_name=f"B2B_Funnel_Plan_{_fname_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.caption("ℹ️ openpyxl nicht installiert")

    _pdf_bytes = generate_pdf(
        revenue_target, deal_size, available_budget, archetype, scenario,
        stage_names, funnel_vals, stage4, deals, total_required, coverage,
        plan_mqls, plan_revenue, ch_calc
    )
    if _pdf_bytes:
        st.download_button(
            "📄 PDF exportieren",
            data=_pdf_bytes,
            file_name=f"B2B_Funnel_Report_{_fname_date}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    else:
        st.caption("ℹ️ fpdf2 nicht installiert")

# ============================================================
# HEADER
# ============================================================

st.title("📊 B2B Revenue Reverse Funnel Planner")
gap_label = f"+€{budget_gap:,.0f}" if budget_gap >= 0 else f"-€{abs(budget_gap):,.0f}"
gap_color = "normal" if budget_gap >= 0 else "inverse"

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Revenue Target",       f"€{revenue_target:,.0f}")
k2.metric("Required Budget",      f"€{total_required:,.0f}", delta=gap_label, delta_color=gap_color)
k3.metric(f"{cov_icon} Coverage", f"{coverage:.0%}")
k4.metric("MQLs needed",          f"{stage4:,.0f}")
k5.metric("Deals needed",         f"{deals:,.0f}")

# ── #9 Budget Status Banner ───────────────────────────────────
if coverage >= 0.9:
    st.success(
        f"✅ **Budget ausreichend** — {coverage:.0%} Deckungsgrad. "
        f"Du hast €{abs(budget_gap):,.0f} Puffer über dem erforderlichen Budget.",
        icon="🟢"
    )
elif coverage >= 0.6:
    st.warning(
        f"⚠️ **Budget-Lücke** — {coverage:.0%} Deckungsgrad. "
        f"Es fehlen €{abs(budget_gap):,.0f}. Prüfe Channel-Mix oder erhöhe das Budget im ⚙️ Inputs-Tab.",
        icon="🟡"
    )
else:
    st.error(
        f"**Kritische Budget-Lücke** — nur {coverage:.0%} Deckungsgrad. "
        f"Es fehlen €{abs(budget_gap):,.0f} ({(1-coverage):.0%} des Bedarfs). "
        f"Empfehlung: Umsatzziel anpassen, Deal Size erhöhen oder Budget aufstocken.",
        icon="🔴"
    )

st.divider()

# ============================================================
# TABS
# ============================================================

tab_inp, tab1, tab2, tab3, tab4, tab5, tab_docs = st.tabs([
    "⚙️ Inputs",
    "🔻 Funnel & Budget",
    "🎲 Risk / Monte Carlo",
    "📅 Monthly Plan",
    "📈 Plan vs. Actual",
    "📡 Channels",
    "📖 Docs",
])

# ══════════════════════════════════════════════════════════════
# TAB: INPUTS
# ══════════════════════════════════════════════════════════════
with tab_inp:

    # ── #26 Onboarding Banner ─────────────────────────────────
    if not st.session_state.get("onboarding_done", False):
        with st.container():
            st.info(
                "👋 **Willkommen im B2B Reverse Funnel Planner!** — "
                "Starte hier mit deinen Zielen (A), wähle deinen Funnel-Typ (B) "
                "und passe die Conversion Rates an (C). "
                "Die Ergebnisse siehst du live in den anderen Tabs. · "
                "*Welcome! Enter your targets in section A, pick a funnel archetype in B, "
                "and adjust conversion rates in C. Results update live across all tabs.*",
                icon="🚀"
            )
            if st.button("✓ Verstanden — nicht mehr anzeigen", key="dismiss_onboarding"):
                st.session_state.onboarding_done = True
                st.rerun()
        st.markdown("")

    # ── A: Business Targets ───────────────────────────────────
    st.markdown('<p class="section-header">A · Business Targets</p>', unsafe_allow_html=True)
    ca1, ca2, ca3 = st.columns(3)
    with ca1:
        st.session_state.inp_revenue = st.number_input(
            "Revenue Target (€)", 10_000, 100_000_000,
            st.session_state.inp_revenue, 50_000, format="%d",
            help="Umsatzziel für die Planungsperiode"
        )
    with ca2:
        st.session_state.inp_deal_size = st.number_input(
            "Avg. Deal Size (€)", 500, 500_000,
            st.session_state.inp_deal_size, 500, format="%d",
            help="Durchschnittlicher Auftragswert"
        )
    with ca3:
        st.session_state.inp_budget = st.number_input(
            "Available Marketing Budget (€)", 0, 20_000_000,
            st.session_state.inp_budget, 10_000, format="%d",
            help="Verfügbares Marketingbudget gesamt"
        )

    st.divider()

    # ── B: Funnel Setup ───────────────────────────────────────
    st.markdown('<p class="section-header">B · Funnel Archetype & Szenario</p>', unsafe_allow_html=True)
    cb1, cb2, cb3 = st.columns([2, 2, 1])

    with cb1:
        prev_arch = st.session_state.inp_archetype
        st.session_state.inp_archetype = st.selectbox(
            "Funnel Archetype",
            list(FUNNEL_ARCHETYPES.keys()),
            index=list(FUNNEL_ARCHETYPES.keys()).index(st.session_state.inp_archetype),
        )
        # Auto-reset CR table when archetype changes
        if st.session_state.inp_archetype != prev_arch:
            st.session_state.inp_cr_df = make_cr_df(st.session_state.inp_archetype)

    with cb2:
        st.session_state.inp_scenario = st.radio(
            "Aktives Szenario",
            ["Worst", "Base", "Best"],
            index=["Worst", "Base", "Best"].index(st.session_state.inp_scenario),
            horizontal=True,
            help="Bestimmt welche Conversion-Rate-Spalte aktiv ist"
        )

    with cb3:
        if st.button("↺ CRs zurücksetzen", use_container_width=True,
                     help="Conversion Rates auf Archetype-Defaults zurücksetzen"):
            st.session_state.inp_cr_df = make_cr_df(st.session_state.inp_archetype)
            st.rerun()

    st.divider()

    # ── C: Conversion Rates ───────────────────────────────────
    st.markdown('<p class="section-header">C · Conversion Rates</p>', unsafe_allow_html=True)
    active_arch = FUNNEL_ARCHETYPES[st.session_state.inp_archetype]
    scen_col_now = {"Worst": "Worst (%)", "Base": "Base (%)", "Best": "Best (%)"}[st.session_state.inp_scenario]
    st.caption(
        f"Aktiv: **{st.session_state.inp_scenario}**-Spalte · "
        f"Archetype: {st.session_state.inp_archetype} · "
        f"Alle Werte in Prozent (z.B. 25 = 25 %)"
    )

    # ── #33 Benchmark-Tooltips per archetype ─────────────────
    _CR_BENCHMARKS = {
        "Classic B2B": [
            "Win Rate · Benchmark B2B: 20–30 %. Enterprise-Deals oft 15–20 %.",
            "SQL→Opp · Benchmark: 35–50 %. Hängt stark von SDR-Qualität ab.",
            "MQL→SQL · Benchmark: 25–40 %. Lead-Scoring verbessert diesen Wert.",
            "Lead→MQL · Benchmark: 10–20 %. Qualifizierungskriterien entscheidend.",
            "Touchpoint→Lead · Benchmark: 1–3 %. Variiert stark je Kanal.",
        ],
        "Enterprise": [
            "Win Rate · Benchmark Enterprise: 15–25 %. Lange Zyklen, weniger Deals.",
            "SQL→Opp · Benchmark: 30–50 %. Hohe Qualifizierung nötig.",
            "MQL→SQL · Benchmark: 45–75 %. Enterprise-MQLs sind oft vorqualifiziert.",
            "Lead→MQL · Benchmark: 15–35 %. ABM erhöht diesen Wert.",
            "Touchpoint→Lead · Benchmark: 0.5–2 %. Nische, hoher Intent nötig.",
        ],
        "SaaS / PLG": [
            "Win Rate · Benchmark SaaS: 18–28 %. PLG-Deals konvertieren schneller.",
            "SQL→Opp · Benchmark: 40–60 %. Product-qualified Leads sind wärmer.",
            "PQL→SQL · Benchmark: 30–55 %. Usage-Signale erhöhen Conversion.",
            "Signup→PQL · Benchmark: 10–25 %. Aktivierungs-Rate ist kritisch.",
            "Visitor→Signup · Benchmark: 2–10 %. Onboarding-UX ist entscheidend.",
        ],
        "Channel / Partner": [
            "Win Rate · Benchmark Channel: 25–40 %. Partner-Deals oft vorgewärmt.",
            "Partner SQL→Opp · Benchmark: 40–70 %. Abhängig von Partner-Enablement.",
            "Partner Lead→SQL · Benchmark: 25–50 %. Co-Selling erhöht Rate.",
            "Contact→Lead · Benchmark: 15–30 %. Partnerqualität entscheidend.",
            "Touchpoint→Contact · Benchmark: 2–7 %. Events & Co-Marketing.",
        ],
    }
    _benchmarks = _CR_BENCHMARKS.get(st.session_state.inp_archetype,
                                     [""] * 5)

    edited_cr = st.data_editor(
        st.session_state.inp_cr_df,
        use_container_width=True,
        num_rows="fixed",
        hide_index=True,
        column_config={
            "Stage":      st.column_config.TextColumn("Funnel-Stufe", disabled=True, width="large"),
            "Worst (%)":  st.column_config.NumberColumn("Worst (%)",  min_value=1, max_value=99, step=1,
                           help="Pessimistischer, aber realistischer Wert — PERT-Gewicht: 1×"),
            "Base (%)":   st.column_config.NumberColumn("Base (%)",   min_value=1, max_value=99, step=1,
                           help="Realistisch erwarteter Wert — PERT-Gewicht: 4× (stärkster Einfluss)"),
            "Best (%)":   st.column_config.NumberColumn("Best (%)",   min_value=1, max_value=99, step=1,
                           help="Optimistischer, aber realistischer Wert — PERT-Gewicht: 1×"),
        },
        key="cr_editor",
    )
    # Show per-stage benchmark hints below table
    with st.expander("📊 Branchen-Benchmarks für diesen Archetype anzeigen", expanded=False):
        for bm in _benchmarks:
            st.caption(f"• {bm}")
    st.session_state.inp_cr_df = edited_cr

    # Visual: aktive CRs als kleine Balken
    active_crs = edited_cr[scen_col_now].tolist()
    fig_cr = go.Figure(go.Bar(
        x=edited_cr["Stage"].tolist(),
        y=active_crs,
        marker_color="#0066cc",
        text=[f"{v}%" for v in active_crs],
        textposition="outside",
    ))
    fig_cr.update_layout(
        height=200, margin=dict(l=0, r=0, t=10, b=0),
        yaxis=dict(range=[0, 105], title="Rate (%)"),
        paper_bgcolor="rgba(0,0,0,0)", showlegend=False,
    )
    st.plotly_chart(fig_cr, use_container_width=True)

    st.divider()

    # ── D: Seasonality ────────────────────────────────────────
    st.markdown('<p class="section-header">D · Saisonalität</p>', unsafe_allow_html=True)
    cd1, cd2 = st.columns([1, 2])

    with cd1:
        prev_season = st.session_state.inp_season
        st.session_state.inp_season = st.selectbox(
            "Seasonality-Profil",
            list(SEASONALITY_PROFILES.keys()),
            index=list(SEASONALITY_PROFILES.keys()).index(st.session_state.inp_season),
        )

        if st.session_state.inp_season == "Eigenes Profil":
            st.caption("Indexwerte pro Monat (werden automatisch normiert):")
            custom_vals = []
            cols_m = st.columns(4)
            for i, month in enumerate(MONTHS):
                default_v = st.session_state.inp_custom_season[i]
                v = cols_m[i % 4].number_input(
                    month, min_value=0.1, max_value=5.0,
                    value=float(default_v), step=0.1,
                    key=f"custom_season_{i}"
                )
                custom_vals.append(v)
            st.session_state.inp_custom_season = custom_vals
            display_weights = custom_vals
        else:
            display_weights = SEASONALITY_PROFILES[st.session_state.inp_season]

    with cd2:
        norm_total = sum(display_weights) or 1
        norm_w = [v / norm_total * 12 for v in display_weights]   # normalized around 1.0
        colors_season = ["#0066cc" if v >= 1.0 else "#80bfff" for v in norm_w]

        fig_season = go.Figure(go.Bar(
            x=MONTHS, y=norm_w,
            marker_color=colors_season,
            text=[f"{v:.2f}x" for v in norm_w],
            textposition="outside",
        ))
        fig_season.add_hline(y=1.0, line_dash="dot", line_color="#6c757d",
                             annotation_text="Ø")
        fig_season.update_layout(
            height=230, margin=dict(l=0, r=0, t=10, b=20),
            yaxis=dict(title="Faktor (1.0 = Durchschnitt)", range=[0, max(norm_w) * 1.3]),
            paper_bgcolor="rgba(0,0,0,0)", showlegend=False,
        )
        st.plotly_chart(fig_season, use_container_width=True)

    st.divider()

    # ── E: Channel Mix & Kosten ───────────────────────────────
    st.markdown('<p class="section-header">E · Channel Mix & Kosten</p>', unsafe_allow_html=True)
    st.caption(
        "**Cost per MQL**: durchschnittliche Kosten, um einen MQL über diesen Kanal zu erzeugen. "
        "**Share**: relativer Anteil — wird automatisch normiert."
    )

    ch_display = st.session_state.channels_df[
        ["group", "activity", "cost_per_mql", "share"]
    ].copy()
    ch_display.columns = ["Kanal-Gruppe", "Aktivität", "Cost per MQL (€)", "Share (%)"]

    edited_ch = st.data_editor(
        ch_display,
        use_container_width=True,
        num_rows="fixed",
        hide_index=True,
        column_config={
            "Kanal-Gruppe": st.column_config.TextColumn(disabled=True),
            "Aktivität":    st.column_config.TextColumn(disabled=True),
            "Cost per MQL (€)": st.column_config.NumberColumn(
                min_value=0, max_value=5_000, step=5,
                help="Kosten pro MQL für diesen Kanal"),
            "Share (%)": st.column_config.NumberColumn(
                min_value=0.0, max_value=100.0, step=0.5,
                help="Relativer MQL-Anteil (wird normiert)"),
        },
        key="inp_ch_editor",
    )
    updated_ch = st.session_state.channels_df.copy()
    updated_ch["cost_per_mql"] = edited_ch["Cost per MQL (€)"].values
    updated_ch["share"]        = edited_ch["Share (%)"].values
    st.session_state.channels_df = updated_ch

    st.divider()

    # ── F: Simulation & Berichtsmonat ─────────────────────────
    st.markdown('<p class="section-header">F · Simulation & Berichtsmonat</p>', unsafe_allow_html=True)
    cf1, cf2 = st.columns(2)
    with cf1:
        st.session_state.inp_n_sims = st.select_slider(
            "Monte Carlo Simulationen",
            [200, 500, 1_000, 2_000],
            value=st.session_state.inp_n_sims,
            help="Mehr Simulationen = präzisere Risikokorridore, aber etwas langsamer"
        )
    with cf2:
        st.session_state.inp_report_month = st.selectbox(
            "Berichtsmonat (für Plan vs. Actual)",
            MONTHS,
            index=MONTHS.index(st.session_state.inp_report_month),
            help="Bis zu welchem Monat wird YTD ausgewertet"
        )

    # ── #37 Inputs Progress Indicator ────────────────────────
    _checks = {
        "A · Business Targets":    st.session_state.inp_revenue != 1_000_000 or st.session_state.inp_deal_size != 5_000,
        "B · Archetype & Szenario": st.session_state.inp_archetype != "Classic B2B" or st.session_state.inp_scenario != "Base",
        "C · Conversion Rates":    not st.session_state.inp_cr_df.equals(make_cr_df(st.session_state.inp_archetype)),
        "D · Saisonalität":        st.session_state.inp_season != "Flat",
        "E · Channel Mix":         True,  # always counts — user has seen it
        "F · Simulation":          st.session_state.inp_n_sims != 500,
    }
    _done  = sum(1 for v in _checks.values() if v)
    _total = len(_checks)
    _pct   = int(_done / _total * 100)
    _bar_color = "#28a745" if _pct == 100 else ("#0066cc" if _pct >= 50 else "#ffc107")

    st.markdown(f"""
<div class="theme-card" style="margin-top:16px;">
  <div style="font-size:0.82rem;font-weight:600;margin-bottom:6px;">
    ⚙️ Konfigurationsfortschritt &nbsp;·&nbsp; {_done}/{_total} Bereiche angepasst
  </div>
  <div style="background:rgba(128,128,128,0.15);border-radius:4px;height:8px;overflow:hidden;">
    <div style="width:{_pct}%;background:{_bar_color};height:8px;
         border-radius:4px;transition:width .3s;"></div>
  </div>
  <div style="margin-top:6px;font-size:0.75rem;opacity:0.7;">
    {"&nbsp; ".join(
        f'<span style="color:{"#28a745" if v else "inherit"};opacity:{"1" if v else "0.45"}">{"✓" if v else "○"} {k}</span>'
        for k, v in _checks.items()
    )}
  </div>
</div>
""", unsafe_allow_html=True)

    st.markdown("")
    st.info(
        "💡 Alle Änderungen hier werden sofort in allen anderen Tabs übernommen.",
        icon="ℹ️"
    )


# ══════════════════════════════════════════════════════════════
# TAB: FUNNEL & BUDGET
# ══════════════════════════════════════════════════════════════
with tab1:
    _settings_bar(archetype, scenario, season_profile, coverage)
    col_funnel, col_budget = st.columns(2)

    with col_funnel:
        st.subheader("Reverse Funnel")
        f_labels = list(reversed(stage_names[1:]))
        f_values = list(reversed(funnel_vals[1:]))
        blue_gradient = ["#003580", "#004ea6", "#0066cc", "#1a8cff", "#4da6ff", "#80bfff"]
        fig_funnel = go.Figure(go.Funnel(
            y=f_labels, x=f_values,
            textinfo="value+percent initial",
            textfont=dict(size=12),
            marker=dict(color=blue_gradient),
            connector=dict(line=dict(color="#dee2e6", width=1, dash="dot")),
        ))
        fig_funnel.update_layout(
            height=430, margin=dict(l=0, r=0, t=10, b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_funnel, use_container_width=True)

    with col_budget:
        st.subheader("Budget Analysis")
        bar_color = "#dc3545" if total_required > available_budget else "#28a745"
        fig_bars = go.Figure()
        fig_bars.add_trace(go.Bar(
            x=["Available Budget", "Required Budget"],
            y=[available_budget, total_required],
            marker_color=["#0066cc", bar_color],
            text=[f"€{available_budget:,.0f}", f"€{total_required:,.0f}"],
            textposition="outside", width=0.45,
        ))
        fig_bars.update_layout(
            height=260, margin=dict(l=0, r=0, t=10, b=0),
            paper_bgcolor="rgba(0,0,0,0)",
            yaxis=dict(tickformat=",.0f", title="€"), showlegend=False,
        )
        st.plotly_chart(fig_bars, use_container_width=True)

        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=min(coverage * 100, 150),
            delta={"reference": 100, "suffix": "%", "relative": False},
            number={"suffix": "%", "valueformat": ".0f", "font": {"size": 34}},
            title={"text": "Budget Coverage", "font": {"size": 13}},
            gauge={
                "axis": {"range": [0, 150], "ticksuffix": "%"},
                "bar": {"color": "#0066cc", "thickness": 0.3},
                "steps": [
                    {"range": [0,  60], "color": "#f8d7da"},
                    {"range": [60, 90], "color": "#fff3cd"},
                    {"range": [90, 150], "color": "#d4edda"},
                ],
                "threshold": {"line": {"color": "#333", "width": 2},
                              "thickness": 0.75, "value": 100},
            },
        ))
        fig_gauge.update_layout(
            height=190, margin=dict(l=20, r=20, t=30, b=10),
            paper_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_gauge, use_container_width=True)

    st.subheader("Funnel Numbers")
    rows = []
    for name, val in zip(stage_names, funnel_vals):
        prefix = "€" if name == "Revenue" else ""
        rows.append({"Stage": name, "Annual": f"{prefix}{val:,.0f}",
                     "Per Month": f"{prefix}{val/12:,.1f}"})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# TAB: MONTE CARLO
# ══════════════════════════════════════════════════════════════
with tab2:
    _settings_bar(archetype, scenario, season_profile, coverage)
    st.subheader("Risk Analysis — Monte Carlo")
    st.caption(f"**{n_sims} Simulationen** · PERT-Verteilungen aus Worst/Base/Best · Cost/MQL ±25 %")

    p10_b  = float(np.percentile(mc_df["budget_req"], 10))
    p50_b  = float(np.percentile(mc_df["budget_req"], 50))
    p90_b  = float(np.percentile(mc_df["budget_req"], 90))
    prob_ok = float((mc_df["budget_req"] <= available_budget).mean())

    ca, cb, cc = st.columns(3)
    with ca:
        st.markdown("**Required Budget (€)**")
        st.metric("P10 – optimistisch",   f"€{p10_b:,.0f}")
        st.metric("Median",               f"€{p50_b:,.0f}")
        st.metric("P90 – konservativ",    f"€{float(np.percentile(mc_df['budget_req'], 90)):,.0f}")
    with cb:
        st.markdown(f"**{stage_names[4]} (MQLs)**")
        st.metric("P10", f"{float(np.percentile(mc_df['stage4'], 10)):,.0f}")
        st.metric("P50", f"{float(np.percentile(mc_df['stage4'], 50)):,.0f}")
        st.metric("P90", f"{float(np.percentile(mc_df['stage4'], 90)):,.0f}")
    with cc:
        st.markdown("**Budget-Ausreichend-Wahrscheinlichkeit**")
        st.metric("Wahrscheinlichkeit", f"{prob_ok:.0%}")
        if prob_ok >= 0.70:
            st.success(f"✅ In {prob_ok:.0%} der Szenarien reicht das Budget.")
        elif prob_ok >= 0.30:
            st.warning(f"⚠️ Budget reicht nur in {prob_ok:.0%} der Szenarien.")
        else:
            st.error(f"🔴 Budget zu knapp in {1-prob_ok:.0%} der Szenarien.")

    # ── Histograms ────────────────────────────────────────────
    p10_mql = float(np.percentile(mc_df["stage4"], 10))
    p50_mql = float(np.percentile(mc_df["stage4"], 50))
    p90_mql = float(np.percentile(mc_df["stage4"], 90))

    fig_hist = make_subplots(rows=1, cols=2,
        subplot_titles=[f"Required Budget — Verteilung über {n_sims} Szenarien",
                        f"{stage_names[4]} — Verteilung über {n_sims} Szenarien"],
        horizontal_spacing=0.10)

    # Left: Budget histogram
    fig_hist.add_trace(go.Histogram(x=mc_df["budget_req"], nbinsx=40,
        marker_color="#0066cc", opacity=0.75, name="Budget",
        hovertemplate="Budget: €%{x:,.0f}<br>Anzahl: %{y}"), row=1, col=1)
    fig_hist.add_vline(x=available_budget, line_dash="dash", line_color="#dc3545", line_width=2,
        annotation_text=f"Verfügbar €{available_budget:,.0f}",
        annotation_font_color="#dc3545", annotation_font_size=11, row=1, col=1)
    fig_hist.add_vline(x=p10_b, line_dash="dot", line_color="#28a745", line_width=1.5,
        annotation_text=f"P10 €{p10_b:,.0f}",
        annotation_font_color="#28a745", annotation_font_size=10,
        annotation_position="top left", row=1, col=1)
    fig_hist.add_vline(x=p50_b, line_dash="dot", line_color="#0066cc", line_width=2,
        annotation_text=f"P50 €{p50_b:,.0f}",
        annotation_font_color="#0066cc", annotation_font_size=11, row=1, col=1)
    fig_hist.add_vline(x=p90_b, line_dash="dot", line_color="#fd7e14", line_width=1.5,
        annotation_text=f"P90 €{p90_b:,.0f}",
        annotation_font_color="#fd7e14", annotation_font_size=10,
        annotation_position="top right", row=1, col=1)

    # Right: MQL histogram
    fig_hist.add_trace(go.Histogram(x=mc_df["stage4"], nbinsx=40,
        marker_color="#28a745", opacity=0.75, name="MQLs",
        hovertemplate="MQLs: %{x:,.0f}<br>Anzahl: %{y}"), row=1, col=2)
    fig_hist.add_vline(x=p10_mql, line_dash="dot", line_color="#28a745", line_width=1.5,
        annotation_text=f"P10 {p10_mql:,.0f}",
        annotation_font_color="#28a745", annotation_font_size=10,
        annotation_position="top left", row=1, col=2)
    fig_hist.add_vline(x=p50_mql, line_dash="dot", line_color="#0066cc", line_width=2,
        annotation_text=f"P50 {p50_mql:,.0f}",
        annotation_font_color="#0066cc", annotation_font_size=11, row=1, col=2)
    fig_hist.add_vline(x=p90_mql, line_dash="dot", line_color="#fd7e14", line_width=1.5,
        annotation_text=f"P90 {p90_mql:,.0f}",
        annotation_font_color="#fd7e14", annotation_font_size=10,
        annotation_position="top right", row=1, col=2)

    fig_hist.update_layout(height=360, showlegend=False,
        paper_bgcolor="rgba(0,0,0,0)", margin=dict(l=0, r=0, t=50, b=0))
    st.plotly_chart(fig_hist, use_container_width=True)

    # ── Risiko-Korridor — normalisiert ───────────────────────
    st.subheader("Risiko-Korridor")
    st.caption(
        "Jede Metrik hat ihre eigene Skala (0 % = P10, 100 % = P90). "
        "Ein **enger Korridor** bedeutet hohe Planungssicherheit — ein **breiter** hohes Risiko. "
        "🟢 P10 = optimistisch · 🔵 P50 = wahrscheinlichster Wert · 🟠 P90 = konservative Planungsbasis"
    )

    corridor_items = [
        ("Deals",                      "deals"),
        (f"{stage_names[5]} (Leads)",  "stage5"),
        (f"{stage_names[4]} (MQLs)",   "stage4"),
        ("Required Budget (€)",        "budget_req"),
    ]

    fig_corridor = go.Figure()
    labels_order = []
    legend_added = False

    for label, col_name in corridor_items:
        p10c = float(np.percentile(mc_df[col_name], 10))
        p50c = float(np.percentile(mc_df[col_name], 50))
        p90c = float(np.percentile(mc_df[col_name], 90))
        labels_order.append(label)

        # Normalise: P10 → 0 %, P90 → 100 %
        spread = (p90c - p10c) if p90c > p10c else 1
        p50_norm = (p50c - p10c) / spread * 100   # position of P50 within corridor

        # Uncertainty label: spread as % of P50
        uncert_pct = spread / p50c * 100 if p50c > 0 else 0
        uncert_label = (
            "🟢 eng"   if uncert_pct < 30 else
            "🟡 mittel" if uncert_pct < 70 else
            "🔴 breit"
        )

        show_leg = not legend_added

        # Gray background 0 → 100
        fig_corridor.add_trace(go.Scatter(
            x=[0, 100], y=[label, label], mode="lines",
            line=dict(color="#e9ecef", width=16),
            showlegend=False, hoverinfo="skip",
        ))
        # Green zone 0 → P50_norm
        fig_corridor.add_trace(go.Scatter(
            x=[0, p50_norm], y=[label, label], mode="lines",
            line=dict(color="#28a745", width=16), opacity=0.5,
            name="← P10 bis P50 (optimistisch)", showlegend=show_leg,
            hoverinfo="skip",
        ))
        # Orange zone P50_norm → 100
        fig_corridor.add_trace(go.Scatter(
            x=[p50_norm, 100], y=[label, label], mode="lines",
            line=dict(color="#fd7e14", width=16), opacity=0.5,
            name="P50 bis P90 → (konservativ)", showlegend=show_leg,
            hoverinfo="skip",
        ))
        legend_added = True

        # P10 marker + actual value below
        fig_corridor.add_trace(go.Scatter(
            x=[0], y=[label], mode="markers+text",
            marker=dict(size=12, color="#28a745", symbol="circle",
                        line=dict(color="white", width=2)),
            text=[f"{p10c:,.0f}"],
            textposition="bottom center",
            textfont=dict(size=9, color="#28a745"),
            showlegend=False,
            hovertemplate=f"<b>P10 — {label}</b><br>Wert: {p10c:,.0f}<extra></extra>",
        ))
        # P50 diamond + actual value above
        fig_corridor.add_trace(go.Scatter(
            x=[p50_norm], y=[label], mode="markers+text",
            marker=dict(size=18, color="#0066cc", symbol="diamond",
                        line=dict(color="white", width=2)),
            text=[f"{p50c:,.0f}"],
            textposition="top center",
            textfont=dict(size=10, color="#0066cc", family="Arial Black"),
            showlegend=False,
            hovertemplate=f"<b>P50 — {label}</b><br>Wert: {p50c:,.0f}<br>Position im Korridor: {p50_norm:.0f} %<extra></extra>",
        ))
        # P90 marker + actual value below
        fig_corridor.add_trace(go.Scatter(
            x=[100], y=[label], mode="markers+text",
            marker=dict(size=12, color="#fd7e14", symbol="circle",
                        line=dict(color="white", width=2)),
            text=[f"{p90c:,.0f}"],
            textposition="bottom center",
            textfont=dict(size=9, color="#fd7e14"),
            showlegend=False,
            hovertemplate=f"<b>P90 — {label}</b><br>Wert: {p90c:,.0f}<extra></extra>",
        ))
        # Uncertainty badge — right-aligned annotation
        fig_corridor.add_annotation(
            x=106, y=label,
            text=f"<b>{uncert_label}</b> ({uncert_pct:.0f} %)",
            showarrow=False,
            font=dict(size=10, color="#555"),
            xanchor="left",
        )

    fig_corridor.update_layout(
        height=320,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=140, t=10, b=30),
        legend=dict(orientation="h", y=1.10, x=0,
                    font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(
            range=[-3, 106],
            tickvals=[0, 25, 50, 75, 100],
            ticktext=["P10<br>(0 %)", "25 %", "P50<br>(Median)", "75 %", "P90<br>(100 %)"],
            showgrid=True, gridcolor="#f0f0f0", zeroline=False,
            tickfont=dict(size=10, color="#888"),
        ),
        yaxis=dict(
            categoryorder="array",
            categoryarray=labels_order,
            tickfont=dict(size=12, color="#333"),
        ),
    )
    st.plotly_chart(fig_corridor, use_container_width=True)


# ══════════════════════════════════════════════════════════════
# TAB: MONTHLY PLAN
# ══════════════════════════════════════════════════════════════
with tab3:
    _settings_bar(archetype, scenario, season_profile, coverage)
    st.subheader(f"Monthly Plan — {season_profile}")

    monthly_data = {
        stage_names[4]:   plan_mqls,
        stage_names[1]:   plan_deals,
        "Revenue (€)":    plan_revenue,
        "Budget (€)":     plan_budget,
    }
    metric_colors = {stage_names[4]: "#0066cc", stage_names[1]: "#28a745",
                     "Revenue (€)": "#fd7e14", "Budget (€)": "#dc3545"}

    fig_monthly = make_subplots(rows=2, cols=2,
        subplot_titles=list(monthly_data.keys()),
        vertical_spacing=0.18, horizontal_spacing=0.08)
    for i, (metric, values) in enumerate(monthly_data.items()):
        row, col = i // 2 + 1, i % 2 + 1
        avg = sum(values) / 12
        fig_monthly.add_trace(go.Bar(x=MONTHS, y=values,
            marker_color=metric_colors[metric], opacity=0.85, showlegend=False,
            hovertemplate=f"%{{x}}: %{{y:,.0f}}"), row=row, col=col)
        fig_monthly.add_hline(y=avg, line_dash="dot", line_color="#6c757d",
            annotation_text=f"⌀ {avg:,.0f}", annotation_position="top right",
            row=row, col=col)
    fig_monthly.update_layout(height=520, paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=0, r=60, t=40, b=0))
    st.plotly_chart(fig_monthly, use_container_width=True)

    table_rows = {"Monat": MONTHS}
    for metric, values in monthly_data.items():
        if "€" in metric:
            table_rows[metric] = [f"€{v:,.0f}" for v in values]
        elif metric == stage_names[1]:
            table_rows[metric] = [f"{v:.1f}" for v in values]
        else:
            table_rows[metric] = [f"{v:,.0f}" for v in values]
    st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# TAB: PLAN VS. ACTUAL
# ══════════════════════════════════════════════════════════════
with tab4:
    _settings_bar(archetype, scenario, season_profile, coverage)
    st.subheader("Plan vs. Actual")
    st.caption("Trage deine monatlichen Ist-Werte ein. Oder lade Musterdaten zum Ausprobieren.")

    report_idx = MONTHS.index(st.session_state.inp_report_month)
    # Musterdaten: realistisches B2B-Jahr
    # Jan–Mär: langsamer Start, Apr–Jun: Normalisierung, Jul–Aug: Sommerloch,
    # Sep–Nov: starker Herbst, Dez: Jahresendspurt (Ø ≈ 100 % des Plans)
    SAMPLE_MULT = [0.88, 1.05, 1.16, 0.91, 0.97, 1.06,
                   0.80, 0.76, 1.14, 1.20, 1.10, 0.97]

    # ── #32 Empty state ───────────────────────────────────────
    _actual_filled = st.session_state.actual_df["Actual MQLs"].sum() > 0
    if not _actual_filled:
        st.markdown(
            """<div class="theme-warn">
            📋 <b>Noch keine Ist-Daten vorhanden.</b>
            Trage deine monatlichen Werte direkt in die Tabelle ein —
            oder klicke auf <b>📥 Musterdaten laden</b> für einen realistischen Jahresverlauf
            (inkl. Sommerdip und Q4-Stärke) zum Ausprobieren.
            <br><small style="opacity:0.65;">No actuals yet. Enter your monthly figures in the table below,
            or load sample data to explore the feature.</small>
            </div>""",
            unsafe_allow_html=True,
        )

    ctrl1, ctrl2, ctrl3 = st.columns([2, 1, 1])
    with ctrl1:
        st.info(f"Berichtsmonat: **{st.session_state.inp_report_month}** (änderbar im Inputs-Tab → Abschnitt F)")
    with ctrl2:
        if st.button("📥 Musterdaten laden", use_container_width=True):
            st.session_state.actual_df = pd.DataFrame({
                "Month":              MONTHS,
                "Actual MQLs":        [round(plan_mqls[i]    * SAMPLE_MULT[i], 0) for i in range(12)],
                "Actual Deals":       [round(plan_deals[i]   * SAMPLE_MULT[i], 1) for i in range(12)],
                "Actual Revenue (€)": [round(plan_revenue[i] * SAMPLE_MULT[i], 0) for i in range(12)],
                "Actual Budget (€)":  [round(plan_budget[i]  * SAMPLE_MULT[i], 0) for i in range(12)],
            })
            st.session_state.inp_report_month = "Dec"
            st.rerun()
    with ctrl3:
        if st.button("🗑️ Daten löschen", use_container_width=True):
            st.session_state.actual_df = pd.DataFrame({
                "Month":              MONTHS,
                "Actual MQLs":        [0.0] * 12,
                "Actual Deals":       [0.0] * 12,
                "Actual Revenue (€)": [0.0] * 12,
                "Actual Budget (€)":  [0.0] * 12,
            })
            st.rerun()

    edited_actual = st.data_editor(
        st.session_state.actual_df, use_container_width=True,
        num_rows="fixed", hide_index=True,
        column_config={
            "Month":              st.column_config.TextColumn("Monat", disabled=True),
            "Actual MQLs":        st.column_config.NumberColumn("Ist MQLs",        min_value=0, step=1,    format="%.0f"),
            "Actual Deals":       st.column_config.NumberColumn("Ist Deals",       min_value=0, step=0.5,  format="%.1f"),
            "Actual Revenue (€)": st.column_config.NumberColumn("Ist Revenue (€)", min_value=0, step=1000, format="€%.0f"),
            "Actual Budget (€)":  st.column_config.NumberColumn("Ist Budget (€)",  min_value=0, step=500,  format="€%.0f"),
        },
        key="actual_editor",
    )
    st.session_state.actual_df = edited_actual

    act_mqls    = edited_actual["Actual MQLs"].tolist()
    act_deals   = edited_actual["Actual Deals"].tolist()
    act_revenue = edited_actual["Actual Revenue (€)"].tolist()
    act_budget  = edited_actual["Actual Budget (€)"].tolist()

    st.divider()

    def ytd_rate(actuals, plans, end_idx):
        a = sum(actuals[:end_idx + 1])
        p = sum(plans[:end_idx + 1])
        return (a / p) if p > 0 else 0.0

    def ytd_icon(rate):
        return "🟢" if rate >= 0.95 else ("🟡" if rate >= 0.75 else "🔴")

    r_mql  = ytd_rate(act_mqls,    plan_mqls,    report_idx)
    r_deal = ytd_rate(act_deals,   plan_deals,   report_idx)
    r_rev  = ytd_rate(act_revenue, plan_revenue, report_idx)
    r_bud  = ytd_rate(act_budget,  plan_budget,  report_idx)

    st.markdown(f"**YTD-Zielerreichung bis {st.session_state.inp_report_month}**")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric(f"{ytd_icon(r_mql)}  MQLs",         f"{r_mql:.0%}",
              delta=f"{sum(act_mqls[:report_idx+1]):,.0f} / {sum(plan_mqls[:report_idx+1]):,.0f}")
    k2.metric(f"{ytd_icon(r_deal)} Deals",         f"{r_deal:.0%}",
              delta=f"{sum(act_deals[:report_idx+1]):.1f} / {sum(plan_deals[:report_idx+1]):.1f}")
    k3.metric(f"{ytd_icon(r_rev)}  Revenue",       f"{r_rev:.0%}",
              delta=f"€{sum(act_revenue[:report_idx+1]):,.0f} / €{sum(plan_revenue[:report_idx+1]):,.0f}")
    k4.metric(f"{ytd_icon(r_bud)}  Budget Spend",  f"{r_bud:.0%}",
              delta=f"€{sum(act_budget[:report_idx+1]):,.0f} / €{sum(plan_budget[:report_idx+1]):,.0f}")

    st.divider()

    metrics_pva = [
        ("MQLs",        plan_mqls,    act_mqls,    "#0066cc", "#80bfff"),
        ("Deals",       plan_deals,   act_deals,   "#28a745", "#a3d9b1"),
        ("Revenue (€)", plan_revenue, act_revenue, "#fd7e14", "#ffd0a0"),
        ("Budget (€)",  plan_budget,  act_budget,  "#6f42c1", "#c9b8f0"),
    ]
    fig_pva = make_subplots(rows=2, cols=2,
        subplot_titles=[m[0] for m in metrics_pva],
        vertical_spacing=0.18, horizontal_spacing=0.08)
    for i, (label, plan_vals, act_vals, color_plan, color_act) in enumerate(metrics_pva):
        row, col = i // 2 + 1, i % 2 + 1
        act_visible = [v if j <= report_idx else None for j, v in enumerate(act_vals)]
        fig_pva.add_trace(go.Bar(x=MONTHS, y=plan_vals, name="Plan",
            marker_color=color_act, opacity=0.5, showlegend=(i == 0),
            legendgroup="plan"), row=row, col=col)
        fig_pva.add_trace(go.Bar(x=MONTHS, y=act_visible, name="Ist",
            marker_color=color_plan, opacity=0.9, showlegend=(i == 0),
            legendgroup="actual"), row=row, col=col)
        fig_pva.add_vline(x=st.session_state.inp_report_month,
            line_dash="dot", line_color="#6c757d", row=row, col=col)
    fig_pva.update_layout(barmode="overlay", height=540,
        paper_bgcolor="rgba(0,0,0,0)", margin=dict(l=0, r=0, t=40, b=0),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    st.plotly_chart(fig_pva, use_container_width=True)

    st.subheader("Kumulativer Revenue-Verlauf")
    cum_plan = [sum(plan_revenue[:i+1]) for i in range(12)]
    cum_act  = [sum(act_revenue[:i+1]) if i <= report_idx else None for i in range(12)]
    fig_cum = go.Figure()
    fig_cum.add_trace(go.Scatter(x=MONTHS, y=cum_plan, name="Plan kumuliert",
        line=dict(color="#adb5bd", width=2, dash="dot"), mode="lines+markers",
        marker=dict(size=6)))
    fig_cum.add_trace(go.Scatter(x=MONTHS[:report_idx+1], y=cum_act[:report_idx+1],
        name="Ist kumuliert", line=dict(color="#fd7e14", width=3),
        mode="lines+markers", marker=dict(size=8),
        fill="tonexty", fillcolor="rgba(253,126,20,0.08)"))
    fig_cum.add_hline(y=revenue_target, line_dash="dash", line_color="#0066cc",
        annotation_text=f"Jahresziel €{revenue_target:,.0f}",
        annotation_font_color="#0066cc")
    fig_cum.update_layout(height=300, paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=0, r=0, t=20, b=0),
        yaxis=dict(tickformat=",.0f", title="Revenue (€)"),
        legend=dict(orientation="h"))
    st.plotly_chart(fig_cum, use_container_width=True)


# ══════════════════════════════════════════════════════════════
# TAB: CHANNELS (Ergebnis-Ansicht)
# ══════════════════════════════════════════════════════════════
with tab5:
    _settings_bar(archetype, scenario, season_profile, coverage)
    st.subheader("Channel-Ergebnisse")
    st.caption("Inputs (Cost per MQL, Share) änderst du im **⚙️ Inputs**-Tab → Abschnitt E.")

    ch_result = calc_channel_budget(stage4, st.session_state.channels_df)
    grp = (ch_result.groupby("group", as_index=False)
           .agg(planned_mqls=("planned_mqls","sum"), required_spend=("required_spend","sum"))
           .sort_values("required_spend", ascending=False))

    col_pie, col_bar = st.columns(2)
    with col_pie:
        fig_pie = px.pie(grp, values="required_spend", names="group", hole=0.42,
            color_discrete_sequence=px.colors.qualitative.Set2)
        fig_pie.update_traces(textposition="outside", textinfo="percent+label")
        fig_pie.update_layout(height=370, paper_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0, r=0, t=10, b=0), showlegend=False)
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_bar:
        fig_bar = go.Figure(go.Bar(
            y=grp["group"], x=grp["required_spend"], orientation="h",
            marker_color="#0066cc",
            text=[f"€{v:,.0f}" for v in grp["required_spend"]],
            textposition="outside"))
        fig_bar.update_layout(height=370, paper_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0, r=90, t=10, b=0),
            xaxis=dict(title="Required Spend (€)", tickformat=",.0f"),
            showlegend=False)
        st.plotly_chart(fig_bar, use_container_width=True)

    grp_d = grp.rename(columns={"group":"Kanal-Gruppe",
                                  "planned_mqls":"Geplante MQLs",
                                  "required_spend":"Required Spend (€)"})
    grp_d["Geplante MQLs"]      = grp_d["Geplante MQLs"].apply(lambda x: f"{x:,.0f}")
    grp_d["Required Spend (€)"] = grp_d["Required Spend (€)"].apply(lambda x: f"€{x:,.0f}")
    st.dataframe(grp_d, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("Details pro Kanal")
    ch_detail = ch_result[["group","activity","planned_mqls","required_spend"]].copy()
    ch_detail.columns = ["Gruppe","Aktivität","Geplante MQLs","Spend (€)"]
    ch_detail["Geplante MQLs"] = ch_detail["Geplante MQLs"].apply(lambda x: f"{x:,.0f}")
    ch_detail["Spend (€)"]     = ch_detail["Spend (€)"].apply(lambda x: f"€{x:,.0f}")
    st.dataframe(ch_detail, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# TAB: DOCS
# ══════════════════════════════════════════════════════════════
with tab_docs:

    st.markdown("""
<style>
.docs-section { margin-bottom: 8px; }
.docs-de { background:#f0f6ff; border-left:3px solid #0066cc;
           padding:10px 14px; border-radius:4px; margin-bottom:6px; }
.docs-en { background:#f6f6f6; border-left:3px solid #888;
           padding:10px 14px; border-radius:4px; margin-bottom:12px; }
.docs-label-de { font-size:0.72rem; font-weight:700; color:#0066cc;
                 text-transform:uppercase; letter-spacing:.05em; }
.docs-label-en { font-size:0.72rem; font-weight:700; color:#888;
                 text-transform:uppercase; letter-spacing:.05em; }
</style>
""", unsafe_allow_html=True)

    st.title("📖 Dokumentation / Documentation")
    st.caption("B2B Revenue Reverse Funnel Planner · © Marko Gross")
    st.divider()

    # ── Quick Start ───────────────────────────────────────────
    with st.expander("🚀 Quick Start — In 3 Schritten / In 3 Steps", expanded=True):
        c_de, c_en = st.columns(2)
        with c_de:
            st.markdown('<p class="docs-label-de">🇩🇪 Deutsch</p>', unsafe_allow_html=True)
            st.markdown("""
**1. ⚙️ Inputs öffnen**
Umsatzziel, Ø Deal Size und verfügbares Marketingbudget eingeben.

**2. Funnel-Archetype wählen**
Classic B2B, Enterprise, SaaS/PLG oder Channel/Partner — passend zu eurem Vertriebsmodell. Die Conversion Rates werden automatisch vorbelegt.

**3. Tabs durchklicken**
Die App berechnet sofort: Wie viele MQLs, Deals und wie viel Budget ihr braucht — inklusive Risikobewertung, Monatsplan und Kanal-Allokation.
""")
        with c_en:
            st.markdown('<p class="docs-label-en">🇬🇧 English</p>', unsafe_allow_html=True)
            st.markdown("""
**1. Open ⚙️ Inputs**
Enter your revenue target, average deal size, and available marketing budget.

**2. Choose a funnel archetype**
Classic B2B, Enterprise, SaaS/PLG, or Channel/Partner — matching your sales motion. Conversion rates are pre-filled automatically.

**3. Explore the tabs**
The app instantly calculates how many MQLs, deals, and how much budget you need — including risk assessment, monthly plan, and channel allocation.
""")

    # ── Methodology ───────────────────────────────────────────
    with st.expander("🔬 Methodik / Methodology — PERT + Monte Carlo"):
        c_de, c_en = st.columns(2)
        with c_de:
            st.markdown('<p class="docs-label-de">🇩🇪 Deutsch</p>', unsafe_allow_html=True)
            st.markdown("""
**PERT-Schätzung**

Statt einer einzigen Conversion-Rate-Annahme arbeitet das Tool mit drei Szenarien pro Funnel-Stufe: Worst Case, Base Case und Best Case. Der gewichtete Mittelwert folgt der PERT-Formel:

> *Ø = (Worst + 4 × Base + Best) / 6*

Der realistischste Fall (Base) fließt mit dem stärksten Gewicht ein — Extremwerte beeinflussen das Ergebnis, dominieren es aber nicht.

---

**Monte Carlo Simulation**

Die Risikoanalyse simuliert bis zu 2.000 mögliche Jahresverläufe. Jede Simulation zieht zufällige Conversion Rates (normalverteilt um den PERT-Mittelwert) und berechnet den resultierenden Budget- und MQL-Bedarf.

Das Ergebnis ist eine **Verteilung statt eines Punktwertes**:
- **P10** = optimistisches Szenario
- **P50** = realistischer Median
- **P90** = konservative Planung
""")
        with c_en:
            st.markdown('<p class="docs-label-en">🇬🇧 English</p>', unsafe_allow_html=True)
            st.markdown("""
**PERT Estimation**

Instead of a single conversion rate, the tool uses three scenarios per funnel stage: Worst, Base, and Best case. The weighted average follows the PERT formula:

> *Mean = (Worst + 4 × Base + Best) / 6*

The most realistic case (Base) carries the strongest weight — extreme values influence the result without dominating it.

---

**Monte Carlo Simulation**

The risk analysis simulates up to 2,000 possible year scenarios. Each run draws random conversion rates (normally distributed around the PERT mean) and calculates the resulting budget and MQL requirements.

The result is a **distribution, not a single number**:
- **P10** = optimistic scenario
- **P50** = realistic median
- **P90** = conservative planning base
""")

    # ── Tab Guide ─────────────────────────────────────────────
    with st.expander("🗂️ Tab-Guide"):
        st.markdown("""
| Tab | DE — Was du siehst & wofür | EN — What you see & why |
|-----|---------------------------|-------------------------|
| ⚙️ **Inputs** | Alle Eingabefelder: Ziele, Archetype, CRs, Saisonalität, Kanäle | All input fields: targets, archetype, CRs, seasonality, channels |
| 🔻 **Funnel & Budget** | Trichter-Grafik, Budget-Gauge, Kanal-Donut — Gesamtbild auf einen Blick | Funnel chart, budget gauge, channel donut — full picture at a glance |
| 🎲 **Risk / Monte Carlo** | Histogramm, P10/P50/P90 — Wie sicher ist der Plan? | Histogram, P10/P50/P90 — how robust is the plan? |
| 📅 **Monthly Plan** | Monatstabelle + saisonaler Verlauf — Wann brauche ich was? | Monthly table + seasonal curve — when do you need what? |
| 📈 **Plan vs. Actual** | Soll-Ist-Vergleich — Abweichungen tracken | Plan vs. actual comparison — track variance month by month |
| 📡 **Channels** | Spend-Verteilung pro Kanal — Budget-Allokation optimieren | Spend breakdown per channel — optimize budget allocation |
| 📖 **Docs** | Diese Dokumentation / This documentation | Diese Dokumentation / This documentation |
""")
        st.caption("💡 DE: Sidebar aufklappen für kompakte KPIs + Excel/PDF-Export. · EN: Open the sidebar for compact KPIs and Excel/PDF export.")

    # ── FAQ ───────────────────────────────────────────────────
    with st.expander("❓ FAQ / Tipps & Tricks"):
        faqs = [
            (
                'Was bedeutet "Budget Coverage"?',
                'Verhältnis von verfügbarem zu benötigtem Budget. 🟢 ≥ 90 % = gut gedeckt, 🟡 60–90 % = Lücke vorhanden, 🔴 < 60 % = kritisch.',
                'What does "Budget Coverage" mean?',
                'Ratio of available to required budget. 🟢 ≥ 90% = well covered, 🟡 60–90% = gap exists, 🔴 < 60% = critical.',
            ),
            (
                'Welchen Funnel-Archetype soll ich wählen?',
                '**Classic B2B** für Standardvertrieb mit Inside/Field Sales. **Enterprise** wenn Zyklen > 6 Monate. **SaaS/PLG** wenn ein Free-Tier oder Trial existiert. **Channel** wenn Partner den Großteil der Pipeline generieren.',
                'Which funnel archetype should I choose?',
                '**Classic B2B** for standard inside/field sales. **Enterprise** for cycles >6 months. **SaaS/PLG** if you have a free tier or trial. **Channel** if partners generate most of your pipeline.',
            ),
            (
                'Wie nutze ich "Plan vs. Actual" am besten?',
                'Monatlich die Ist-Zahlen eintragen — die App zeigt sofort, ob ihr auf Kurs seid. Mit "Musterdaten laden" bekommt ihr einen realistischen Jahresverlauf als Startpunkt.',
                'How do I get the most out of "Plan vs. Actual"?',
                'Enter actuals monthly — the app instantly shows whether you\'re on track. Use "Load sample data" for a realistic seasonal starting point.',
            ),
            (
                'Kann ich den Plan exportieren?',
                'Ja — Sidebar aufklappen (Pfeil links oben). Dort findest du **📥 Excel exportieren** (4 Sheets: Summary, Monthly Plan, Plan vs. Actual, Channels) und **📄 PDF exportieren** (kompakter 1-Seiter).',
                'Can I export the plan?',
                'Yes — open the sidebar (arrow top left). You\'ll find **📥 Excel Export** (4 sheets: Summary, Monthly Plan, Plan vs. Actual, Channels) and **📄 PDF Export** (compact 1-pager).',
            ),
            (
                'Wie viele Monte-Carlo-Simulationen sind sinnvoll?',
                '500 ist ein guter Standard. Für Präsentationen oder finale Planungen empfehlen sich 1.000–2.000 Runs für stabilere Perzentile. Mehr Runs = längere Ladezeit.',
                'How many Monte Carlo simulations make sense?',
                '500 is a good default. For presentations or final planning, 1,000–2,000 runs give more stable percentiles. More runs = longer load time.',
            ),
        ]
        for q_de, a_de, q_en, a_en in faqs:
            c_de, c_en = st.columns(2)
            with c_de:
                st.markdown(f'<p class="docs-label-de">🇩🇪</p>', unsafe_allow_html=True)
                st.markdown(f"**{q_de}**")
                st.markdown(f"→ {a_de}")
            with c_en:
                st.markdown(f'<p class="docs-label-en">🇬🇧</p>', unsafe_allow_html=True)
                st.markdown(f"**{q_en}**")
                st.markdown(f"→ {a_en}")
            st.divider()

    # ── Glossary ──────────────────────────────────────────────
    with st.expander("📚 Glossar / Glossary"):
        glossary = [
            ("Reverse Funnel",    "Vom Ziel rückwärts rechnen: Umsatz → Deals → MQLs → Budget",            "Working backwards from goal: Revenue → Deals → MQLs → Budget"),
            ("PERT",              "Program Evaluation and Review Technique — gewichtete Drei-Punkt-Schätzung (Worst / Base / Best)", "Weighted three-point estimation method (Worst / Base / Best)"),
            ("Monte Carlo",       "Statistische Simulation mit bis zu 2.000 zufälligen Szenarien zur Risikoabschätzung",             "Statistical simulation using up to 2,000 random scenarios to quantify risk"),
            ("MQL",               "Marketing Qualified Lead — Lead, der Marketingkriterien erfüllt",         "Lead that meets marketing qualification criteria"),
            ("SQL",               "Sales Qualified Lead — vom Vertrieb akzeptierter Lead",                   "Lead accepted and validated by the sales team"),
            ("Opportunity (Opp)", "Qualifizierte Verkaufschance im CRM",                                     "Qualified sales opportunity tracked in the CRM"),
            ("Win Rate",          "Anteil gewonnener Deals an allen Opportunities",                          "Share of deals won vs. total opportunities"),
            ("Conversion Rate",   "Übergangsrate zwischen zwei aufeinanderfolgenden Funnel-Stufen",           "Transition rate between two consecutive funnel stages"),
            ("P10 / P50 / P90",   "Perzentile: P10 = optimistisch, P50 = Median, P90 = konservativ",        "Percentiles: P10 = optimistic, P50 = median, P90 = conservative"),
            ("Budget Coverage",   "Deckungsgrad: Verfügbares ÷ benötigtes Budget",                          "Coverage ratio: Available ÷ required budget"),
            ("Cost per MQL",      "Durchschnittliche Kosten pro MQL je Kanal",                               "Average cost to generate one MQL per channel"),
            ("Touchpoint",        "Erster Kontaktpunkt eines Leads mit Marketing-Maßnahmen",                 "First point of contact between a lead and a marketing activity"),
            ("Saisonalität",      "Monatliche Gewichtung des Jahresplans (z. B. Q4-lastig)",                 "Monthly weighting of the annual plan (e.g. Q4-heavy)"),
            ("Archetype",         "Vordefiniertes Funnel-Muster passend zum Vertriebsmodell",                "Pre-defined funnel pattern matching your sales motion"),
        ]
        g_de, g_en = st.columns(2)
        with g_de:
            st.markdown('<p class="docs-label-de">🇩🇪 Deutsch</p>', unsafe_allow_html=True)
            rows_de = "| Begriff | Erklärung |\n|---|---|\n"
            for term, de, _ in glossary:
                rows_de += f"| **{term}** | {de} |\n"
            st.markdown(rows_de)
        with g_en:
            st.markdown('<p class="docs-label-en">🇬🇧 English</p>', unsafe_allow_html=True)
            rows_en = "| Term | Explanation |\n|---|---|\n"
            for term, _, en in glossary:
                rows_en += f"| **{term}** | {en} |\n"
            st.markdown(rows_en)

    # ── Power Features ────────────────────────────────────────
    with st.expander("✨ Power Features & Hidden Gems"):
        st.markdown("*DE: Kleine Dinge, die einen großen Unterschied machen. · EN: Small things that make a big difference.*")
        st.markdown("")

        features = [
            (
                "🎥", "Screencast aufnehmen",
                "Drei-Punkte-Menü oben rechts (⋮) → **Record a screencast** — direkt im Browser, kein Extra-Tool nötig. Perfekt für Demo-Videos oder kurze Erklärungen für Kollegen.",
                "Record a screencast",
                "Three-dot menu top right (⋮) → **Record a screencast** — right in the browser, no extra tool needed. Perfect for demo videos or quick walkthroughs for colleagues.",
            ),
            (
                "📥", "Excel & PDF Export",
                "Sidebar aufklappen (Pfeil links) → **📥 Excel exportieren** (4 Sheets) oder **📄 PDF exportieren** (1-seitiger Report). Dateiname enthält automatisch das Tagesdatum.",
                "Excel & PDF Export",
                "Open the sidebar (arrow on the left) → **📥 Excel Export** (4 sheets) or **📄 PDF Export** (1-page report). Filename automatically includes today's date.",
            ),
            (
                "📊", "Interaktive Charts",
                "Alle Charts sind vollständig interaktiv: Zoomen, Panning, Hover-Tooltips mit genauen Werten, Legenden ein-/ausblenden. Kamera-Icon in der Chart-Toolbar → Chart als **PNG herunterladen**.",
                "Interactive Charts",
                "All charts are fully interactive: zoom, pan, hover tooltips with exact values, toggle legend entries. Camera icon in the chart toolbar → download chart as **PNG**.",
            ),
            (
                "📋", "Tabellen als CSV",
                "Über jede Datentabelle hovern → kleines Download-Icon erscheint oben rechts → direkt als **CSV exportieren**, ohne Extra-Button.",
                "Tables as CSV",
                "Hover over any data table → a small download icon appears top right → export directly as **CSV**, no extra button needed.",
            ),
            (
                "🌙", "Dark Mode",
                "Drei-Punkte-Menü (⋮) → **Settings** → Theme umschalten zwischen Hell und Dunkel.",
                "Dark Mode",
                "Three-dot menu (⋮) → **Settings** → toggle between light and dark theme.",
            ),
            (
                "↔️", "Sidebar ein-/ausklappen",
                "Pfeil am linken Rand — gibt mehr Platz für Charts. Die KPIs (Revenue, Budget, Coverage) bleiben jederzeit erreichbar.",
                "Collapse / expand sidebar",
                "Arrow on the left edge — gives more space for charts. KPIs (Revenue, Budget, Coverage) remain accessible at any time.",
            ),
            (
                "🔗", "Direktlink teilen",
                "Die Streamlit-Cloud-URL ist sofort teilbar — **kein Login nötig** für Empfänger. Link kopieren und an Stakeholder schicken.",
                "Share a direct link",
                "The Streamlit Cloud URL is instantly shareable — **no login required** for recipients. Copy and send to stakeholders.",
            ),
            (
                "📱", "Mobile-tauglich",
                "Die App läuft auch im Smartphone-Browser — gut für schnelle Checks unterwegs oder Präsentationen vom Tablet.",
                "Mobile-friendly",
                "The app runs in any mobile browser — great for quick checks on the go or tablet presentations.",
            ),
            (
                "🔄", "Live-Berechnung",
                "Jede Eingabe löst sofort eine Neuberechnung aus — kein 'Berechnen'-Button nötig. Änderungen in CRs oder Budget sind in Echtzeit in allen Tabs sichtbar.",
                "Live recalculation",
                "Every input instantly triggers a full recalculation — no 'Calculate' button needed. Changes to CRs or budget are reflected across all tabs in real time.",
            ),
            (
                "📂", "Musterdaten laden",
                "Im Tab **📈 Plan vs. Actual** → Button **'Musterdaten laden'** — füllt ein realistisches Jahresprofil (inkl. Sommerdip, Q4-Stärke) als Startpunkt.",
                "Load sample data",
                "In the **📈 Plan vs. Actual** tab → **'Load sample data'** button — fills a realistic seasonal year profile (summer dip, Q4 strength) as a starting point.",
            ),
        ]

        for icon, title_de, text_de, title_en, text_en in features:
            c_de, c_en = st.columns(2)
            with c_de:
                st.markdown(f'<p class="docs-label-de">🇩🇪 {icon} {title_de}</p>', unsafe_allow_html=True)
                st.markdown(text_de)
            with c_en:
                st.markdown(f'<p class="docs-label-en">🇬🇧 {icon} {title_en}</p>', unsafe_allow_html=True)
                st.markdown(text_en)
            st.markdown("")

    # ── Use Cases ─────────────────────────────────────────────
    with st.expander("💼 Use Cases — Marketing Budget Planning"):
        st.markdown("---")
        uc1_de, uc1_en = st.columns(2)
        with uc1_de:
            st.markdown('<p class="docs-label-de">🇩🇪 Use Case 1: Jahresplanung Marketing-Budget</p>', unsafe_allow_html=True)
            st.markdown("""
**Szenario:** Ein B2B-SaaS-Unternehmen plant sein Marketingbudget für das kommende Jahr. Umsatzziel: 2 Mio. €, Ø Deal Size: 8.000 €.

**Vorgehen:**
1. Archetype **SaaS / PLG** wählen, Umsatzziel auf 2.000.000 € setzen
2. Base-Case CRs aus CRM-Historien eintragen (Worst/Best als ±30 %)
3. Monte Carlo auf 1.000 Simulationen — **P90 als Planungsbasis** für den Budget-Request
4. Saisonalität **Q4 heavy** → zeigt, dass ~38 % des Budgets in Q3/Q4 fließen müssen
5. **Excel exportieren** → direkt als Anlage für den CFO-Review

**Ergebnis:** Statt einer einzigen Zahl liefert die Präsentation drei Szenarien (P10/P50/P90) mit Wahrscheinlichkeiten — deutlich überzeugender als ein einzelner Bauchgefühl-Wert.
""")
        with uc1_en:
            st.markdown('<p class="docs-label-en">🇬🇧 Use Case 1: Annual Marketing Budget Planning</p>', unsafe_allow_html=True)
            st.markdown("""
**Scenario:** A B2B SaaS company plans its marketing budget for the upcoming year. Revenue target: EUR 2M, avg. deal size: EUR 8,000.

**Steps:**
1. Select archetype **SaaS / PLG**, set revenue target to 2,000,000
2. Enter Base Case CRs from CRM history (Worst/Best as ±30%)
3. Run Monte Carlo at 1,000 simulations — use **P90 as planning base** for budget request
4. Seasonality **Q4 heavy** → shows ~38% of budget must land in Q3/Q4
5. **Export to Excel** → attach directly to CFO budget review

**Result:** Instead of a single number, the presentation delivers three scenarios (P10/P50/P90) with probabilities — far more compelling than a single gut-feeling estimate.
""")

        st.markdown("---")
        uc2_de, uc2_en = st.columns(2)
        with uc2_de:
            st.markdown('<p class="docs-label-de">🇩🇪 Use Case 2: Mid-Year Review — Ist das Ziel noch erreichbar?</p>', unsafe_allow_html=True)
            st.markdown("""
**Szenario:** Q2 ist abgeschlossen. Das Team hat 45 % der geplanten MQLs erreicht, aber nur 30 % des Umsatzziels. Der VP fragt: Reichen die verbleibenden 6 Monate noch?

**Vorgehen:**
1. Tab **📈 Plan vs. Actual** öffnen → Ist-Zahlen für Jan–Jun eintragen
2. Lücke sofort sichtbar: Wo weicht es am stärksten ab — MQLs, Win Rate oder Deal Size?
3. Im Tab **⚙️ Inputs** die CRs für H2 anpassen (z. B. höhere Win Rate durch neues Sales-Enablement)
4. Monte Carlo zeigt: Mit optimierten CRs liegt P50 noch knapp über Ziel — P90 erfordert +15 % Budget
5. **PDF exportieren** → kompakter 1-Seiter für das Management-Meeting

**Ergebnis:** Datenbasierte Entscheidung statt Schätzung — entweder Jahresziel anpassen oder Budget nachschießen, mit klarer quantitativer Begründung.
""")
        with uc2_en:
            st.markdown('<p class="docs-label-en">🇬🇧 Use Case 2: Mid-Year Review — Is the target still achievable?</p>', unsafe_allow_html=True)
            st.markdown("""
**Scenario:** Q2 is done. The team hit 45% of planned MQLs but only 30% of the revenue target. The VP asks: Can we still close the gap in H2?

**Steps:**
1. Open tab **📈 Plan vs. Actual** → enter actuals for Jan–Jun
2. Gap immediately visible: where is the biggest deviation — MQLs, win rate, or deal size?
3. In **⚙️ Inputs**, adjust H2 CRs (e.g. higher win rate from new sales enablement)
4. Monte Carlo shows: with optimized CRs, P50 is just above target — P90 requires +15% budget
5. **Export to PDF** → compact 1-pager for the management meeting

**Result:** Data-driven decision instead of guesswork — either adjust the annual target or inject budget, with a clear quantitative rationale.
""")


# ============================================================
# FOOTER
# ============================================================

st.divider()
st.caption(
    "**B2B Revenue Reverse Funnel Planner** · "
    "Methodology: PERT Estimation + Monte Carlo Simulation · "
    "Built with Streamlit & Plotly · "
    "© Marko Gross"
)
